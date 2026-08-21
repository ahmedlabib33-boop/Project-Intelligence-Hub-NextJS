from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import re
import shutil
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from advanced_analytics import build_advanced_analytics
from project_chart_payloads import build_project_chart_payloads
from project_report_artifacts import ensure_project_report_artifacts
from universal_report_engine_adapter import (
    FAMILY_MANIFEST_NAME,
    ensure_universal_report_engine_catalog,
    is_released_artifact,
)

# Canonical inputs embed lossless JSON snapshots in ``payload_json``.  Letter
# registers can exceed Python's small default CSV field limit (131,072 bytes).
# Raise the parser limit once for every controlled project input.
try:
    csv.field_size_limit(sys.maxsize)
except OverflowError:
    csv.field_size_limit(2**31 - 1)


ROOT = Path(__file__).resolve().parents[1]
# This delivery is self-contained.  Generated website data must always come
# from this workspace, never from a second local checkout or an environment
# override that could publish stale project data.
CANONICAL_ROOT = ROOT
PROJECTS_ROOT = CANONICAL_ROOT / "projects"
SOURCE_OUTPUTS_ROOT = CANONICAL_ROOT / "11-outputs"
OUTPUTS_ROOT = ROOT / "11-outputs"
WEBSITE_PUBLIC = ROOT / "website" / "public"
DATA_ROOT = WEBSITE_PUBLIC / "data"
GENERATED_ROOT = WEBSITE_PUBLIC / "generated"
WEBSITE_SOURCE_GENERATED = ROOT / "website" / "src" / "generated"

# Project workspaces show source-backed, paginated previews. Keeping every raw row
# in browser payloads makes selected-project navigation unreliable on mobile and
# does not add report capability; complete report artifacts stay in Output Studio.
WORKSPACE_TABLE_ROW_LIMIT = 200


def slugify(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9]+", "-", value.strip()).strip("-")
    return value or "project"


def normalize_field_name(value: Any) -> str:
    """Match CSV headers despite spaces, punctuation, case, or underscores."""
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def safe_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null", "n/a", "na", "-"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    numeric_match = re.search(r"[-+]?\d+(?:\.\d+)?", text.replace(",", ""))
    if numeric_match is None:
        return None
    try:
        number = float(numeric_match.group(0))
    except ValueError:
        return None
    if not math.isfinite(number):
        return None
    return -number if negative else number


def safe_percent(value: Any) -> float | None:
    number = safe_float(value)
    if number is None:
        return None
    if number > 1:
        number = number / 100.0
    return max(0.0, min(number, 1.0))


def read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception:
        return {}


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    for encoding in ("utf-8-sig", "utf-8", "cp1256", "cp1252"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.DictReader(handle))
        except Exception:
            continue
    return []


CANONICAL_BUNDLE_NAMES = (
    "01_project_contract.csv",
    "02_schedule_activities.csv",
    "03_schedule_logic.csv",
    "04_progress_evm.csv",
    "05_milestones_scurve.csv",
    "06_delay_events.csv",
    "07_tia_evidence_scenarios.csv",
    "08_commercial_payments_claims.csv",
    "09_risks_rfi_interfaces.csv",
    "10_letters_intelligence.csv",
)

LOGICAL_DATASETS = (
    "projects", "contracts", "payments", "planned_cash_flow", "progress", "evm",
    "risks", "claims", "activities", "milestones", "delay_events", "s_curve",
    "wbs", "historical_outcomes",
)


def _bundle_value(value: Any) -> str:
    return "" if value is None else str(value)


def _decode_canonical_input_bundles(data_dir: Path) -> tuple[dict[str, list[dict[str, str]]], dict[str, Path]]:
    """Rebuild legacy logical rows from the ten compact, lossless input bundles.

    The bundles preserve original headers and row order in payload_json.  This is
    intentionally in-memory: the ten files remain the only active project inputs.
    """
    tables: dict[str, dict[str, Any]] = {}
    sources: dict[str, Path] = {}
    found_bundle = False
    for bundle_name in CANONICAL_BUNDLE_NAMES:
        bundle_path = data_dir / bundle_name
        for record in read_csv_rows(bundle_path):
            # Readable canonical CSVs expose one source row per spreadsheet row.
            # They deliberately replace the former payload_json envelope while
            # retaining the same logical source tables and values.
            readable_table = str(record.get("source_table") or "").strip()
            readable_path = str(record.get("source_path") or "").strip()
            if readable_table and readable_path:
                found_bundle = True
                table = tables.setdefault(readable_table, {"headers": [], "rows": [], "direct_rows": []})
                try:
                    row_order = int(str(record.get("row_order") or "0"))
                except ValueError:
                    row_order = len(table["direct_rows"])
                direct_row = {
                    key: value
                    for key, value in record.items()
                    if key not in {"source_table", "source_path", "row_order"} and value not in (None, "")
                }
                table["direct_rows"].append((row_order, direct_row))
                sources.setdefault(readable_table, bundle_path)
                continue
            if not record.get("bundle_version") or not record.get("source_file"):
                continue
            found_bundle = True
            source_name = Path(str(record["source_file"])).stem
            table = tables.setdefault(source_name, {"headers": [], "rows": []})
            sources.setdefault(source_name, bundle_path)
            try:
                payload = json.loads(str(record.get("payload_json") or "null"))
            except json.JSONDecodeError:
                continue
            kind = str(record.get("row_kind") or "").strip().casefold()
            if kind == "schema" and isinstance(payload, dict):
                headers = payload.get("headers")
                if isinstance(headers, list):
                    table["headers"] = [str(header) for header in headers]
            elif kind == "data" and isinstance(payload, list):
                try:
                    row_order = int(str(record.get("row_order") or "0"))
                except ValueError:
                    row_order = len(table["rows"])
                table["rows"].append((row_order, payload))
    if not found_bundle:
        return {}, {}

    result: dict[str, list[dict[str, str]]] = {}
    for name, table in tables.items():
        direct_rows = table.get("direct_rows", [])
        if direct_rows:
            result[name] = [row for _, row in sorted(direct_rows, key=lambda item: item[0])]
            continue
        headers = table["headers"]
        result[name] = [
            {header: _bundle_value(values[index]) if index < len(values) else "" for index, header in enumerate(headers)}
            for _, values in sorted(table["rows"], key=lambda item: item[0])
            if headers
        ]

    # Some original TIA files carry their numbered filename in source_file.
    # Alias them to the dashboard's stable logical dataset names.
    for original_name, logical_name in {
        "08- payments": "payments",
    }.items():
        if original_name in result and logical_name not in result:
            result[logical_name] = result[original_name]
            sources[logical_name] = sources[original_name]
    # activity_master is the lossless merge of the former activity, EVM, and
    # progress feeds. Reconstruct each old logical view only in memory.
    master_rows = result.get("activity_master", [])
    for target, prefix in (("activities", "activity__"), ("evm", "evm__"), ("progress", "progress__")):
        rebuilt: list[dict[str, str]] = []
        for master in master_rows:
            row = {column[len(prefix):]: value for column, value in master.items() if column.startswith(prefix)}
            if str(row.get("activity_id") or "").strip():
                rebuilt.append(row)
        if rebuilt:
            result[target] = rebuilt
            sources[target] = sources.get("activity_master", data_dir / "02_schedule_activities.csv")
    return result, sources


def load_project_input_rows(data_dir: Path) -> tuple[dict[str, list[dict[str, str]]], dict[str, Path]]:
    """Return the logical datasets used by the app from canonical or legacy inputs."""
    decoded, sources = _decode_canonical_input_bundles(data_dir)
    rows: dict[str, list[dict[str, str]]] = {}
    source_paths: dict[str, Path] = {}
    for dataset in LOGICAL_DATASETS:
        canonical_name = dataset
        if decoded:
            rows[dataset] = list(decoded.get(canonical_name, []))
            if canonical_name in sources:
                source_paths[dataset] = sources[canonical_name]
        else:
            filename = "progress_updates.csv" if dataset == "progress" else f"{dataset}.csv"
            path = data_dir / filename
            rows[dataset] = read_csv_rows(path)
            source_paths[dataset] = path
    return rows, source_paths


def load_canonical_letters_payload(project: dict[str, Any], data_dir: Path) -> dict[str, Any] | None:
    """Read the project-scoped Letters snapshot from canonical input 10.

    The ten-input redesign stores the former workbook and inbox result as one
    lossless snapshot row.  It must be read directly instead of looking for the
    archived ``07-letters_intelligence`` workbook folder.
    """
    bundle_path = data_dir / "10_letters_intelligence.csv"
    project_id = str(project.get("project_id") or "").strip()
    if not bundle_path.exists() or not project_id:
        return None

    records = read_csv_rows(bundle_path)
    if records and "record_type" in records[0] and "sheet_name" in records[0]:
        inbox_files: list[dict[str, Any]] = []
        sheets: dict[str, dict[str, Any]] = {}
        letters_metadata = {"record_type", "sheet_name", "row_order", "file_name", "relative_path", "extension", "size_kb", "modified"}
        for record in records:
            record_type = str(record.get("record_type") or "").strip().casefold()
            if record_type == "inbox_file":
                inbox_files.append(
                    {
                        "name": str(record.get("file_name") or "Letter file"),
                        "relative_path": str(record.get("relative_path") or record.get("file_name") or ""),
                        "extension": str(record.get("extension") or "file"),
                        "size_kb": safe_float(record.get("size_kb")) or 0,
                        "modified": str(record.get("modified") or ""),
                    }
                )
                continue
            if record_type != "letter_register":
                continue
            row_project_id = str(record.get("project_id") or "").strip()
            if row_project_id and row_project_id != project_id:
                continue
            sheet_name = str(record.get("sheet_name") or "Letters Register")
            sheet = sheets.setdefault(sheet_name, {"name": sheet_name, "columns": [], "rows": []})
            row = {
                key: excel_value(value)
                for key, value in record.items()
                if key not in letters_metadata and value not in (None, "")
            }
            row["project_id"] = project_id
            for column in row:
                if column not in sheet["columns"]:
                    sheet["columns"].append(column)
            sheet["rows"].append(row)
        published_sheets = [
            {
                "name": sheet["name"],
                "row_count": len(sheet["rows"]),
                "column_count": len(sheet["columns"]),
                "columns": sheet["columns"],
                "rows": sheet["rows"][:WORKSPACE_TABLE_ROW_LIMIT],
                "truncated": len(sheet["rows"]) > WORKSPACE_TABLE_ROW_LIMIT,
            }
            for sheet in sheets.values()
        ]
        return {
            "folder": "01-data/import_templates",
            "inbox_files": inbox_files,
            "inbox_file_count": len(inbox_files),
            "workbook": {"file": bundle_path.name, "exists": True, "source_workbook": "letters_intelligence.xlsx", "sheets": published_sheets},
            "workbook_tables": {
                "file": bundle_path.name,
                "exists": True,
                "sheets": published_sheets,
                "rejected_cross_project_rows": 0,
                "source_scope": "selected_project_only",
                "inbox_auto_ingest": True,
                "source_mode": "readable_csv_register",
            },
            "source_file": bundle_path.name,
        }

    for record in records:
        if str(record.get("row_kind") or "").strip().casefold() != "snapshot":
            continue
        try:
            snapshot = json.loads(str(record.get("payload_json") or "null"))
        except json.JSONDecodeError:
            continue
        if not isinstance(snapshot, dict):
            continue

        workbook_tables = snapshot.get("workbook_tables")
        if not isinstance(workbook_tables, dict):
            continue

        published_sheets: list[dict[str, Any]] = []
        rejected_rows = 0
        for source_sheet in workbook_tables.get("sheets", []):
            if not isinstance(source_sheet, dict):
                continue
            source_rows = source_sheet.get("rows")
            if not isinstance(source_rows, list):
                source_rows = []
            scoped_rows: list[dict[str, Any]] = []
            for source_row in source_rows:
                if not isinstance(source_row, dict):
                    continue
                row = {str(key): excel_value(value) for key, value in source_row.items()}
                row_project_id = str(row.get("project_id") or "").strip()
                if row_project_id and row_project_id != project_id:
                    rejected_rows += 1
                    continue
                row["project_id"] = project_id
                scoped_rows.append(row)
            columns = [str(column) for column in source_sheet.get("columns", [])]
            if "project_id" not in columns:
                columns.append("project_id")
            declared_number = safe_float(source_sheet.get("row_count"))
            declared_count = int(declared_number) if declared_number is not None else len(scoped_rows)
            published_sheets.append(
                {
                    "name": str(source_sheet.get("name") or "Letters Register"),
                    "row_count": declared_count if declared_count >= len(scoped_rows) else len(scoped_rows),
                    "column_count": len(columns),
                    "columns": columns,
                    "rows": scoped_rows[:WORKSPACE_TABLE_ROW_LIMIT],
                    "truncated": bool(source_sheet.get("truncated")) or len(scoped_rows) > WORKSPACE_TABLE_ROW_LIMIT,
                }
            )

        source_files = snapshot.get("inbox_files")
        inbox_files = [
            {
                "name": str(item.get("name") or "Letter file"),
                "relative_path": str(item.get("relative_path") or item.get("name") or ""),
                "extension": str(item.get("extension") or "file"),
                "size_kb": safe_float(item.get("size_kb")) or 0,
                "modified": str(item.get("modified") or ""),
            }
            for item in source_files
            if isinstance(item, dict)
        ] if isinstance(source_files, list) else []

        workbook = snapshot.get("workbook") if isinstance(snapshot.get("workbook"), dict) else {}
        return {
            "folder": "01-data/import_templates",
            "inbox_files": inbox_files,
            "inbox_file_count": len(inbox_files),
            "workbook": {
                "file": bundle_path.name,
                "exists": True,
                "source_workbook": str(workbook.get("file") or "letters_intelligence.xlsx"),
                "sheets": published_sheets,
            },
            "workbook_tables": {
                "file": bundle_path.name,
                "exists": True,
                "sheets": published_sheets,
                "rejected_cross_project_rows": rejected_rows,
                "source_scope": "selected_project_only",
                "inbox_auto_ingest": True,
                "source_mode": "canonical_csv_snapshot",
            },
            "source_file": bundle_path.name,
        }
    return None


def preview_rows_table(rows: list[dict[str, Any]], source_path: Path, limit: int = 8) -> dict[str, Any]:
    columns = list(rows[0].keys()) if rows else []
    return {
        "file": source_path.name,
        "exists": source_path.exists(),
        "row_count": len(rows),
        "column_count": len(columns),
        "columns": columns,
        "rows": rows[:limit],
    }


def workspace_rows_table(rows: list[dict[str, Any]], source_path: Path, limit: int = WORKSPACE_TABLE_ROW_LIMIT) -> dict[str, Any]:
    columns = list(rows[0].keys()) if rows else []
    return {
        "file": source_path.name,
        "exists": source_path.exists(),
        "row_count": len(rows),
        "column_count": len(columns),
        "columns": columns,
        "rows": rows[:limit],
        "truncated": len(rows) > limit,
        "source_path": source_path.name,
    }

def compact_file_record(path: Path, base: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "name": path.name,
        "relative_path": path.relative_to(base).as_posix(),
        "extension": path.suffix.lower().lstrip(".") or "file",
        "size_kb": round(stat.st_size / 1024, 1),
        "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
    }


def list_project_files(path: Path, base: Path, limit: int = 120) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    files = [
        compact_file_record(child, base)
        for child in sorted(path.rglob("*"))
        if child.is_file() and not child.name.startswith(".")
    ]
    return files[:limit]


def preview_table(path: Path, limit: int = 8) -> dict[str, Any]:
    rows = read_csv_rows(path)
    columns: list[str] = []
    if rows:
        columns = list(rows[0].keys())
    elif path.exists():
        try:
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle)
                columns = next(reader, [])
        except Exception:
            columns = []
    return {
        "file": path.name,
        "exists": path.exists(),
        "row_count": len(rows),
        "column_count": len(columns),
        "columns": columns,
        "rows": rows[:limit],
    }


def workspace_table(path: Path, limit: int = WORKSPACE_TABLE_ROW_LIMIT) -> dict[str, Any]:
    """Return a project-scoped table suitable for the digital workspace.

    The portfolio remains a compact executive payload. Full rows travel only in
    the selected project's JSON file so one project's source evidence cannot be
    rendered in another project's workspace.
    """
    rows = read_csv_rows(path)
    columns = list(rows[0].keys()) if rows else []
    return {
        "file": path.name,
        "exists": path.exists(),
        "row_count": len(rows),
        "column_count": len(columns),
        "columns": columns,
        "rows": rows[:limit],
        "truncated": len(rows) > limit,
        "source_path": path.name,
    }


def xlsx_summary(path: Path, limit: int = 8) -> dict[str, Any]:
    summary = {"file": path.name, "exists": path.exists(), "sheets": []}
    if not path.exists():
        return summary
    try:
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows_iter = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows_iter, [])]
            preview_rows = []
            row_count = 0
            for row in rows_iter:
                row_count += 1
                if len(preview_rows) < limit:
                    preview_rows.append({headers[idx] if idx < len(headers) and headers[idx] else f"Column {idx + 1}": value for idx, value in enumerate(row)})
            summary["sheets"].append(
                {
                    "name": sheet_name,
                    "row_count": row_count,
                    "column_count": len(headers),
                    "columns": headers,
                    "rows": preview_rows,
                }
            )
    except Exception as exc:
        summary["error"] = str(exc)
    return summary


def xlsx_workspace_tables(path: Path, limit_per_sheet: int = WORKSPACE_TABLE_ROW_LIMIT) -> dict[str, Any]:
    """Expose full project-owned workbook sheets with an explicit safety cap."""
    result: dict[str, Any] = {"file": path.name, "exists": path.exists(), "sheets": []}
    if not path.exists():
        return result
    try:
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows_iter = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows_iter, [])]
            headers = [header or f"Column {index + 1}" for index, header in enumerate(headers)]
            records: list[dict[str, Any]] = []
            row_count = 0
            for row in rows_iter:
                row_count += 1
                if len(records) < limit_per_sheet:
                    records.append({headers[index]: excel_value(value) for index, value in enumerate(row[:len(headers)])})
            result["sheets"].append(
                {
                    "name": sheet_name,
                    "row_count": row_count,
                    "column_count": len(headers),
                    "columns": headers,
                    "rows": records,
                    "truncated": row_count > limit_per_sheet,
                }
            )
    except Exception as exc:
        result["error"] = str(exc)
    return result


def _project_scoped_letter_frame(frame: Any, project_id: str) -> tuple[Any, int]:
    """Keep only rows belonging to the active project before publishing letters."""
    import pandas as pd

    scoped = frame.copy().where(frame.notna(), "") if isinstance(frame, pd.DataFrame) else pd.DataFrame()
    project_column = next(
        (column for column in scoped.columns if normalize_field_name(column) == "projectid"),
        None,
    )
    rejected = 0
    if project_column is not None:
        normalized = scoped[project_column].astype(str).str.strip()
        allowed = normalized.eq("") | normalized.eq(project_id)
        rejected = int((~allowed).sum())
        scoped = scoped.loc[allowed].copy()
    scoped["project_id"] = project_id
    return scoped, rejected


def _frames_to_workspace_tables(
    sheets: dict[str, Any], project_id: str, limit_per_sheet: int = WORKSPACE_TABLE_ROW_LIMIT
) -> dict[str, Any]:
    """Convert legacy Streamlit letter tables into selected-project JSON tables."""
    result: dict[str, Any] = {"file": "letters_intelligence.xlsx", "exists": True, "sheets": []}
    rejected_rows = 0
    for sheet_name, frame in sheets.items():
        scoped, rejected = _project_scoped_letter_frame(frame, project_id)
        rejected_rows += rejected
        columns = [str(column) for column in scoped.columns]
        records = [
            {column: excel_value(value) for column, value in row.items()}
            for row in scoped.head(limit_per_sheet).to_dict("records")
        ]
        result["sheets"].append(
            {
                "name": str(sheet_name),
                "row_count": int(len(scoped)),
                "column_count": len(columns),
                "columns": columns,
                "rows": records,
                "truncated": len(scoped) > limit_per_sheet,
            }
        )
    result["rejected_cross_project_rows"] = rejected_rows
    result["source_scope"] = "selected_project_only"
    return result


def merge_canonical_letters_with_inbox(
    project: dict[str, Any], canonical_letters: dict[str, Any], inbox_dir: Path
) -> dict[str, Any]:
    """Overlay the live, project-local inbox on the canonical letters register.

    The ten-input letters CSV remains the historical register authority.  New
    controlled correspondence placed in either direction folder is processed in
    memory by the same Streamlit ingestion engine and then published with the
    selected project's generated JSON.  This deliberately does not rewrite the
    canonical CSV or the original evidence file.
    """
    project_id = str(project.get("project_id") or "").strip()
    workspace = dict(canonical_letters.get("workbook_tables") or {})
    if not project_id or not inbox_dir.exists():
        return workspace
    try:
        import pandas as pd

        source_sheets: dict[str, Any] = {}
        for sheet in workspace.get("sheets") or []:
            if not isinstance(sheet, dict):
                continue
            name = str(sheet.get("name") or "").strip()
            if not name:
                continue
            rows = sheet.get("rows") if isinstance(sheet.get("rows"), list) else []
            source_sheets[name] = pd.DataFrame(rows).fillna("")

        canonical_src = CANONICAL_ROOT / "src"
        if canonical_src.exists() and str(canonical_src) not in sys.path:
            sys.path.insert(0, str(canonical_src))
        if str(CANONICAL_ROOT) not in sys.path:
            sys.path.insert(0, str(CANONICAL_ROOT))
        from construction_system.letters_auto_ingest import merge_inbox_letters
        from contract_claims_center import extract_text_from_path

        merged = _frames_to_workspace_tables(
            merge_inbox_letters(source_sheets, inbox_dir, extract_text_from_path), project_id
        )
        merged["file"] = str(workspace.get("file") or "10_letters_intelligence.csv")
        merged["exists"] = True
        merged["inbox_auto_ingest"] = True
        merged["source_workbook"] = str(workspace.get("source_workbook") or "letters_intelligence.xlsx")
        return merged
    except Exception as exc:
        workspace["inbox_processing_error"] = str(exc)
        return workspace


def build_letters_workspace_tables(
    project: dict[str, Any], workbook_path: Path, inbox_dir: Path
) -> dict[str, Any]:
    """Run the existing Streamlit letters workflow without writing to source files.

    Workbook registers and inbox files are merged in memory through the same
    `merge_inbox_letters` engine that powers the canonical Streamlit workspace.
    The result is then stamped and filtered to the selected `project_id` before
    it is included in that project's Vercel payload.
    """
    project_id = str(project.get("project_id") or "").strip()
    if not project_id:
        return {"file": workbook_path.name, "exists": workbook_path.exists(), "sheets": [], "error": "Project ID is required."}
    try:
        import pandas as pd

        source_sheets: dict[str, Any] = {}
        if workbook_path.exists():
            workbook = pd.ExcelFile(workbook_path)
            source_sheets = {
                str(sheet_name): pd.read_excel(workbook, sheet_name=sheet_name).fillna("")
                for sheet_name in workbook.sheet_names
            }
        canonical_src = CANONICAL_ROOT / "src"
        if canonical_src.exists() and str(canonical_src) not in sys.path:
            sys.path.insert(0, str(canonical_src))
        if str(CANONICAL_ROOT) not in sys.path:
            sys.path.insert(0, str(CANONICAL_ROOT))
        from construction_system.letters_auto_ingest import merge_inbox_letters
        from contract_claims_center import extract_text_from_path

        merged_sheets = merge_inbox_letters(source_sheets, inbox_dir, extract_text_from_path)
        tables = _frames_to_workspace_tables(merged_sheets, project_id)
        tables["file"] = workbook_path.name
        tables["exists"] = workbook_path.exists() or inbox_dir.exists()
        tables["inbox_auto_ingest"] = True
        return tables
    except Exception as exc:
        return {
            "file": workbook_path.name,
            "exists": workbook_path.exists(),
            "sheets": [],
            "error": f"Letters workspace processing failed: {exc}",
            "source_scope": "selected_project_only",
        }


def json_safe_sql_value(value: Any) -> Any:
    if isinstance(value, bytes):
        return f"[binary {len(value)} bytes]"
    return excel_value(value)


def sqlite_table_rows(path: Path, limit_per_table: int = WORKSPACE_TABLE_ROW_LIMIT) -> dict[str, Any]:
    """Read project-local claims records without exposing files from other projects."""
    result: dict[str, Any] = {"exists": path.exists(), "tables": {}, "error": None}
    if not path.exists():
        return result
    try:
        connection = sqlite3.connect(path)
        table_names = [
            str(row[0])
            for row in connection.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
            if not str(row[0]).startswith("sqlite_")
        ]
        for table_name in table_names:
            quoted_name = table_name.replace('"', '""')
            cursor = connection.execute(f'SELECT * FROM "{quoted_name}" LIMIT ?', (limit_per_table,))
            columns = [str(column[0]) for column in cursor.description or []]
            records = [
                {columns[index]: json_safe_sql_value(value) for index, value in enumerate(row)}
                for row in cursor.fetchall()
            ]
            total = connection.execute(f'SELECT COUNT(*) FROM "{quoted_name}"').fetchone()[0]
            result["tables"][table_name] = {
                "file": f"{path.name}:{table_name}",
                "exists": True,
                "row_count": total,
                "column_count": len(columns),
                "columns": columns,
                "rows": records,
                "truncated": total > limit_per_table,
            }
        connection.close()
    except Exception as exc:
        result["error"] = str(exc)
    return result


def normalize_markdown_text(text: str) -> str:
    return (
        text.replace("â€”", "—")
        .replace("â€“", "–")
        .replace("â€™", "'")
        .replace("â€œ", '"')
        .replace("â€", '"')
    )


def markdown_section(text: str, heading: str) -> str:
    start = text.find(heading)
    if start < 0:
        return ""
    next_heading = re.search(r"\n##\s+", text[start + len(heading):])
    if not next_heading:
        return text[start:]
    return text[start:start + len(heading) + next_heading.start()]


def parse_markdown_table(section: str) -> list[dict[str, str]]:
    lines = [line.strip() for line in section.splitlines() if line.strip().startswith("|")]
    if len(lines) < 3:
        return []
    headers = [cell.strip() for cell in lines[0].strip("|").split("|")]
    rows: list[dict[str, str]] = []
    for line in lines[2:]:
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        if len(cells) != len(headers):
            continue
        rows.append({headers[index]: cells[index] for index in range(len(headers))})
    return rows


def excel_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value


def submitted_tia_guide_root(base: Path, project: dict[str, Any]) -> Path | None:
    candidates = [
        base / "02-delay_analysis" / "TIA submitted Guide",
        base / "02-delay_analysis" / "submitted_tia",
        base / "02-delay_analysis" / "submitted_guide",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate

    # A project may explicitly point to a legacy guide during a staged migration.
    # The path is declared in that project's manifest, never inferred from its name.
    configured_path = str(project.get("submitted_tia_guide_path") or "").strip()
    if configured_path:
        configured = Path(configured_path)
        candidate = configured if configured.is_absolute() else (base / configured)
        try:
            candidate.resolve().relative_to(CANONICAL_ROOT.resolve())
        except ValueError:
            return None
        if candidate.is_dir():
            return candidate
    return None


def submitted_tia_visual_category(name: str) -> str:
    """Classify a submitted exhibit by its topic, without deriving any schedule result."""
    normalized = name.lower().replace("_", " ").replace("-", " ")
    if any(term in normalized for term in ("timeline", "chronology", "finish movement")):
        return "Timeline & Events"
    if any(term in normalized for term in ("float", "critical path", "fragnet", "concurrency", "affected activities")):
        return "Path, Float & Fragnets"
    if any(term in normalized for term in ("ev01", "ev02", "fishbone")):
        return "Event Detail"
    if any(term in normalized for term in ("methodology", "waterfall", "entitlement")):
        return "Methodology & Entitlement"
    return "Submitted Exhibits"


def submitted_tia_visual_key(path: Path) -> str:
    """Collapse alternate source variants while preferring revised and large-font exhibits."""
    stem = path.stem.lower()
    stem = stem.replace("_large_font", "").replace("_inkscape", "")
    return stem.replace("eventnew", "event")


def submitted_tia_visual_priority(path: Path) -> int:
    name = path.stem.lower()
    return (2 if "large_font" in name else 0) + (1 if "eventnew" in name else 0)


def build_submitted_tia_visuals(project: dict[str, Any], base: Path) -> dict[str, Any]:
    """Expose only selected-project submitted TIA exhibits stored with project evidence."""
    visual_root = base / "02-delay_analysis" / "submitted_visuals"
    if not visual_root.exists():
        return {
            "available": False,
            "status": "No submitted visual exhibits detected",
            "scope_note": "Add client or consultant TIA SVG/PNG exhibits under this selected project's 02-delay_analysis/submitted_visuals folder.",
            "evidentiary_note": "No submitted exhibit is available for this selected project.",
            "visuals": [],
        }

    supported = {".svg", ".png"}
    candidates = [
        item for item in sorted(visual_root.rglob("*"))
        if item.is_file() and item.suffix.lower() in supported and not item.name.startswith(".")
    ]
    selected: dict[str, Path] = {}
    for item in candidates:
        key = submitted_tia_visual_key(item)
        current = selected.get(key)
        if current is None or submitted_tia_visual_priority(item) > submitted_tia_visual_priority(current):
            selected[key] = item

    project_slug = slugify(project["project_folder_name"])
    visuals = []
    for item in sorted(selected.values(), key=lambda path: (submitted_tia_visual_category(path.stem), path.name.lower())):
        relative_path = item.relative_to(visual_root).as_posix()
        visuals.append({
            "name": item.name,
            "label": re.sub(r"[_-]+", " ", item.stem).strip(),
            "category": submitted_tia_visual_category(item.stem),
            "relative_path": relative_path,
            "url": f"/generated/{project_slug}/tia-submitted-exhibits/{slugify(item.stem)}{item.suffix.lower()}",
        })

    return {
        "available": bool(visuals),
        "status": "Submitted visual exhibits available" if visuals else "No supported submitted visual exhibits detected",
        "scope_note": "These client-submission exhibits are available only in this selected project's Delay Analysis workspace.",
        "evidentiary_note": "Figures and values within submitted exhibits are source material, not a Vercel recalculation. Confirm EOT, concurrency, and compensation in Primavera P6 and the project evidence record.",
        "visuals": visuals,
    }


def parse_fragnet_comparison(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows: list[dict[str, Any]] = []
        for row in sheet.iter_rows(values_only=True, min_row=4):
            values = list(row[:11])
            if not any(value is not None and str(value).strip() for value in values):
                continue
            if values[0] and not values[2] and not values[9]:
                continue
            rows.append({
                "delay_event": excel_value(values[0]),
                "data_date": excel_value(values[1]),
                "activity_id": excel_value(values[2]),
                "milestone_activity_name": excel_value(values[3]),
                "before_total_float": excel_value(values[4]),
                "before_forecast_finish": excel_value(values[5]),
                "after_total_float": excel_value(values[6]),
                "after_forecast_finish": excel_value(values[7]),
                "float_change_days": excel_value(values[8]),
                "finish_movement_days": excel_value(values[9]),
                "impact_assessment": excel_value(values[10]),
            })
        return rows[:40]
    except Exception:
        return []


def parse_event_register(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["EVT-01"] if "EVT-01" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "").replace("\n", " ").strip() for value in next(iterator, [])]
        rows: list[dict[str, Any]] = []
        for raw in iterator:
            if not any(value is not None and str(value).strip() for value in raw):
                continue
            item = {headers[index]: excel_value(raw[index]) for index in range(min(len(headers), len(raw))) if headers[index]}
            if item.get("Event"):
                rows.append(item)
        return rows[:80]
    except Exception:
        return []


def build_submitted_tia_payload(project: dict[str, Any], base: Path) -> dict[str, Any]:
    guide_root = submitted_tia_guide_root(base, project)
    if not guide_root:
        return {
            "available": False,
            "status": "No submitted TIA guide detected",
            "scope_note": "Generic TIA readiness mode only. Add a project-specific submitted guide under 02-delay_analysis/submitted_tia to activate submitted-event logic.",
        }

    guide_md = guide_root / "THE_BIG_COMPLETE_PACKAGE_TECHNICAL_GUIDE.md"
    text = normalize_markdown_text(guide_md.read_text(encoding="utf-8", errors="ignore")) if guide_md.exists() else ""
    level3 = markdown_section(text, "## 5. Level 3")
    level4 = markdown_section(text, "## 6. Level 4")
    evidence = markdown_section(text, "## 4. Level 2")
    submitted_results = parse_markdown_table(markdown_section(level3, "### 5.3 Submitted-event results"))
    reconciliation = parse_markdown_table(markdown_section(level3, "### 5.4 Current Level 4 reconciliation"))
    evidence_status_controls = parse_markdown_table(markdown_section(evidence, "### 4.3 Evidence status controls"))

    evidence_gap_section = evidence.split("### 4.5 Current critical evidence gaps", 1)[-1] if "### 4.5 Current critical evidence gaps" in evidence else ""
    evidence_gaps = [line.strip("- ").strip() for line in evidence_gap_section.splitlines() if line.strip().startswith("-")][:20]
    warning_section = level3.split("### 5.5 Model-integrity warnings", 1)[-1] if "### 5.5 Model-integrity warnings" in level3 else ""
    model_warnings = [line.strip("- ").strip() for line in warning_section.splitlines() if line.strip().startswith("-")][:20]

    fragnet_rows = parse_fragnet_comparison(guide_root / "TIA Before After Fragnet Comparison.xlsx")
    event_register = parse_event_register(guide_root / "Event Registers.xlsx")
    source_files = list_project_files(guide_root, guide_root, 180)
    event_folders = [
        {
            "name": child.name,
            "file_count": len([item for item in child.rglob("*") if item.is_file()]),
            "xer_count": len(list(child.rglob("*.xer"))),
            "pdf_count": len(list(child.rglob("*.pdf"))),
            "xlsx_count": len(list(child.rglob("*.xlsx"))),
        }
        for child in sorted(guide_root.iterdir())
        if child.is_dir() and child.name[:2].isdigit()
    ]
    visuals = [
        {
            "name": item.name,
            "relative_path": item.relative_to(guide_root).as_posix(),
            "url": f"/generated/{slugify(project['project_folder_name'])}/tia-submitted-guide/{slugify(item.stem)}{item.suffix.lower()}",
        }
        for item in sorted(list(guide_root.glob("*.svg")) + list((guide_root / "New folder").glob("*.svg")) + list((guide_root / "New folder").glob("*.png")))
    ][:18]

    return {
        "available": True,
        "status": "Submitted TIA guide active",
        "guide_folder": str(guide_root),
        "scope_note": "Submitted TIA logic is applied only to this selected project because the guide identifies THE BIG Phase 01.",
        "governance_principle": "A schedule movement is not automatically an entitlement; evidence, contract procedure, causation, schedule integrity, concurrency, mitigation, and determination must be tested separately.",
        "decision_gates": [
            "Authority gate",
            "Clause gate",
            "Evidence gate",
            "Procedure gate",
            "Causation gate",
            "Schedule gate",
            "Concurrency gate",
            "Mitigation gate",
            "Determination gate",
        ],
        "submitted_results": submitted_results,
        "level4_reconciliation": reconciliation,
        "evidence_status_controls": evidence_status_controls,
        "evidence_gaps": evidence_gaps,
        "model_integrity_warnings": model_warnings,
        "fragnet_comparison": fragnet_rows,
        "event_register": event_register,
        "event_folders": event_folders,
        "visuals": visuals,
        "source_files": source_files,
        "recommended_next_moves": [
            "Issue neutral evidence directions without prejudging entitlement.",
            "Require event-specific indexed files for steel, IFC, MEP and RFI events.",
            "Verify native documents and proof of issue/receipt for all relied-upon records.",
            "Reconcile each XER pair to the same approved update, data date and schedule options.",
            "Audit calendars, constraints, open ends, relationships, lags, out-of-sequence settings and non-fragnet changes.",
            "Assess event movement independently before concurrency and prevent double counting.",
            "Separate EOT assessment from prolongation cost and other monetary entitlement.",
        ],
    }


def sqlite_table_counts(path: Path) -> dict[str, Any]:
    result: dict[str, Any] = {"exists": path.exists(), "tables": {}, "error": None}
    if not path.exists():
        return result
    try:
        connection = sqlite3.connect(path)
        cursor = connection.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        for (table_name,) in cursor.fetchall():
            try:
                count = connection.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
            except Exception:
                count = None
            result["tables"][table_name] = count
        connection.close()
    except Exception as exc:
        result["error"] = str(exc)
    return result


def dataframe_workspace_table(frame: Any, file_name: str, limit: int = WORKSPACE_TABLE_ROW_LIMIT) -> dict[str, Any]:
    """Convert a canonical pandas result into the same safe table contract as CSV data."""
    if frame is None:
        return {"file": file_name, "exists": False, "row_count": 0, "column_count": 0, "columns": [], "rows": []}
    try:
        import pandas as pd  # type: ignore

        columns = [str(column) for column in frame.columns]
        rows: list[dict[str, Any]] = []
        for record in frame.head(limit).to_dict(orient="records"):
            cleaned: dict[str, Any] = {}
            for key, value in record.items():
                if pd.isna(value):
                    cleaned[str(key)] = None
                elif hasattr(value, "isoformat"):
                    cleaned[str(key)] = value.isoformat()
                elif isinstance(value, (str, int, float, bool)):
                    cleaned[str(key)] = value
                else:
                    cleaned[str(key)] = str(value)
            rows.append(cleaned)
        return {
            "file": file_name,
            "exists": True,
            "row_count": int(len(frame)),
            "column_count": len(columns),
            "columns": columns,
            "rows": rows,
            "truncated": len(frame) > limit,
            "source_path": file_name,
        }
    except Exception as exc:
        return {"file": file_name, "exists": False, "row_count": 0, "column_count": 0, "columns": [], "rows": [], "error": str(exc)}


def build_legacy_generic_tia_snapshot_archived(delay_dir: Path, project: dict[str, Any] | None = None) -> dict[str, Any]:
    """Run the Streamlit TIA engine for one project and publish only its results.

    A TIA failure is a visible readiness state, never a replacement result from
    another project or a guessed EOT value.
    """
    if not delay_dir.exists():
        return {"status": "missing", "message": "Delay TIA source folder is missing.", "tables": {}}
    try:
        import pandas as pd  # type: ignore

        source_src = CANONICAL_ROOT / "src"
        if str(source_src) not in sys.path:
            sys.path.insert(0, str(source_src))
        from construction_system.steel_delay_tia import SteelTiaSettings, run_steel_delay_tia_analysis
        from construction_system.tia_source_governance import build_tia_source_assessment

        lookup = {re.sub(r"[^a-z0-9]+", "", path.name.lower()): path for path in delay_dir.glob("*.csv")}

        def frame(fragment: str) -> Any:
            path = lookup.get(re.sub(r"[^a-z0-9]+", "", fragment.lower()))
            return pd.read_csv(path, dtype=object) if path and path.exists() else pd.DataFrame()

        event_frames = [frame(name) for name in ("07-ifc_conflict.csv", "08-payments.csv", "09-rfi_status.csv")]
        event_frames = [item for item in event_frames if not item.empty]
        event_df = pd.concat(event_frames, ignore_index=True, sort=False) if event_frames else pd.DataFrame()
        analysis = run_steel_delay_tia_analysis(
            p6_df=frame("04-p6_activity_export.csv"),
            steel_df=frame("03-employer_steel_supply_at_site.csv"),
            requirement_df=frame("02-master_activity_steel_analysis.csv"),
            relationship_df=frame("05-relationship_file.csv"),
            contract_library_df=frame("06-contract_library.csv"),
            delay_events_df=event_df,
            settings=SteelTiaSettings(),
        )
        tables = {
            key: dataframe_workspace_table(value, f"TIA engine - {key}")
            for key, value in analysis.items()
            if hasattr(value, "columns")
        }
        raw_kpis = analysis.get("kpis", {}) if isinstance(analysis.get("kpis"), dict) else {}
        kpis: dict[str, Any] = {}
        for key, value in raw_kpis.items():
            if pd.isna(value):
                kpis[str(key)] = None
            elif hasattr(value, "isoformat"):
                kpis[str(key)] = value.isoformat()
            elif isinstance(value, (str, int, float, bool)):
                kpis[str(key)] = value
            else:
                kpis[str(key)] = str(value)
        governed = build_tia_source_assessment(project) if project else {
            "status": "awaiting_project_context",
            "source_scope": "selected_project_only",
            "event_register": [],
            "summary": {"final_eot_days": None, "final_eot_status": "Not calculated"},
        }
        return {
            "status": "ready",
            "message": "Canonical Streamlit TIA engine completed for the selected project. Results remain indicative until P6 recalculation is verified.",
            "kpis": kpis,
            "tables": tables,
            "source_governance": governed,
        }
    except Exception as exc:
        return {"status": "needs_review", "message": f"Canonical TIA engine could not complete: {exc}", "tables": {}}


def build_controlled_tia_snapshot(project: dict[str, Any]) -> dict[str, Any]:
    """Publish only the selected project's controlled TIA run.

    The former generic CSV engine is intentionally not called here.  The
    controlled adapter creates an unreviewed project-local draft when an
    approved source changes, or a setup state when no project package exists.
    """
    try:
        source_src = CANONICAL_ROOT / "src"
        if str(source_src) not in sys.path:
            sys.path.insert(0, str(source_src))
        from construction_system.controlled_tia import refresh_controlled_tia_run

        return refresh_controlled_tia_run(project)
    except Exception as exc:
        return {
            "engine": "controlled-project-tia",
            "project_id": str(project.get("project_id") or ""),
            "project_key": str(project.get("project_key") or ""),
            "status": "SETUP_REQUIRED",
            "approval_status": "not_submitted",
            "message": f"Controlled TIA setup could not be inspected: {exc}",
            "workflow_tabs": [],
            "source_integrity": {"files": [], "signature": {"status": "not_checked"}},
            "schedule_cpm": {"xer_pairs": []},
            "events_and_fragnets": {"events": [], "event_exhibits": []},
            "concurrency_and_entitlement": {"controls": []},
            "eot_position": {"label": "Not available"},
            "ai_scope": {"status": "guidance_only"},
            "missing_evidence": ["Resolve the controlled TIA setup error before publishing a result."],
            "reconciliation_items": [],
        }


def public_controlled_tia_payload(snapshot: dict[str, Any], public_slug: str) -> dict[str, Any]:
    """Remove workstation paths while retaining evidence file lineage."""
    public = json.loads(json.dumps(snapshot, default=str))
    public.pop("run_path", None)
    integrity = public.get("source_integrity")
    if isinstance(integrity, dict):
        integrity.pop("release_path", None)
        archive = integrity.get("archive")
        if isinstance(archive, dict):
            archive.pop("path", None)
        for item in integrity.get("files", []):
            if isinstance(item, dict):
                item.pop("path", None)
    event_section = public.get("events_and_fragnets")
    if isinstance(event_section, dict):
        exhibits = event_section.get("event_exhibits")
        if isinstance(exhibits, list):
            for exhibit in exhibits:
                if not isinstance(exhibit, dict):
                    continue
                relative_path = str(exhibit.get("source_relative_path") or "").strip()
                if not relative_path:
                    continue
                source_path = Path(relative_path)
                exhibit["url"] = (
                    f"/generated/{public_slug}/tia-controlled-event-exhibits/"
                    f"{slugify(source_path.stem)}{source_path.suffix.lower()}"
                )
                exhibit.pop("sha256", None)
    view_exhibits = public.get("view_exhibits")
    if isinstance(view_exhibits, list):
        for exhibit in view_exhibits:
            if not isinstance(exhibit, dict):
                continue
            relative_path = str(exhibit.get("source_relative_path") or "").strip()
            if not relative_path:
                continue
            source_path = Path(relative_path)
            exhibit["url"] = (
                f"/generated/{public_slug}/tia-controlled-view-exhibits/"
                f"{slugify(source_path.stem)}{source_path.suffix.lower()}"
            )
            exhibit.pop("sha256", None)
    return public


def build_four_pipeline_snapshot(project: dict[str, Any], controlled_tia: dict[str, Any]) -> dict[str, Any]:
    """Publish a compact read-only governance view from the controlled run.

    The former generic four-pipeline evaluator remains historic material.  It
    must not reintroduce generic TIA inputs into the active project payload.
    """
    integrity = controlled_tia.get("source_integrity", {}) if isinstance(controlled_tia, dict) else {}
    return {
        "project_id": str(project.get("project_id") or ""),
        "project_key": str(project.get("project_key") or ""),
        "assessment_profile": "controlled_tia_readiness",
        "assessment_status": str(controlled_tia.get("status") or "SETUP_REQUIRED"),
        "source_scope": "selected_project_only",
        "gates": [
            {"name": "Source integrity", "status": str(integrity.get("signature", {}).get("status") or "not_checked")},
            {"name": "Schedule and CPM", "status": str(controlled_tia.get("schedule_cpm", {}).get("status") or "awaiting_evidence")},
            {"name": "Event and fragnet", "status": str(controlled_tia.get("events_and_fragnets", {}).get("status") or "awaiting_evidence")},
            {"name": "Concurrency and entitlement", "status": str(controlled_tia.get("concurrency_and_entitlement", {}).get("status") or "awaiting_evidence")},
            {"name": "P6 approval", "status": str(controlled_tia.get("approval_status") or "not_submitted")},
        ],
        "missing_actions": list(controlled_tia.get("missing_evidence") or []),
        "pipeline_rows": [],
        "source_inventory": [
            {**item, "project_id": str(project.get("project_id") or ""), "project_key": str(project.get("project_key") or "")}
            for item in (integrity.get("files") or [])
            if isinstance(item, dict)
        ],
        "evidence_ledger": list(controlled_tia.get("events_and_fragnets", {}).get("events") or []),
        "summary": {"eot_position": controlled_tia.get("eot_position", {}).get("label")},
    }


def build_contract_controls_snapshot(
    project: dict[str, Any], contracts_dir: Path, evidence_dir: Path
) -> dict[str, Any]:
    """Refresh and publish selected-project contract/evidence controls only."""

    try:
        if str(CANONICAL_ROOT) not in sys.path:
            sys.path.insert(0, str(CANONICAL_ROOT))
        from contract_claims_center import (  # type: ignore
            build_project_contract_controls,
            persist_contract_analysis,
        )

        db_path = contracts_dir / "contract_claims.db"
        status = persist_contract_analysis(db_path, contracts_dir, rebuild=False)
        controls = build_project_contract_controls(
            db_path=db_path,
            contracts_dir=contracts_dir,
            evidence_dir=evidence_dir,
            project_id=str(project.get("project_id") or ""),
            project_key=str(project.get("project_key") or ""),
        )
        public_status = dict(status) if isinstance(status, dict) else {}
        auto_library_status = public_status.get("auto_library_status")
        if isinstance(auto_library_status, dict):
            public_status["auto_library_status"] = {
                key: value
                for key, value in auto_library_status.items()
                if key not in {"library_path", "database_path", "contracts_dir", "evidence_dir"}
            }
        return {"status": public_status, "controls": controls}
    except Exception as exc:
        return {
            "status": {"knowledge_base_status": "Needs review", "error": str(exc)},
            "controls": {
                "project_id": str(project.get("project_id") or ""),
                "project_key": str(project.get("project_key") or ""),
                "source_scope": "selected_project_only",
                "clause_controls": [],
                "evidence_ledger": [],
            },
        }


def build_feature_payload(project: dict[str, Any], rows: dict[str, list[dict[str, str]]], source_paths: dict[str, Path]) -> dict[str, Any]:
    base = Path(project["path"])
    data_dir = base / "01-data" / "import_templates"
    delay_dir = base / "02-delay_analysis" / "unified_tia_csv"
    schedule_dir = base / "03-schedule"
    contracts_dir = base / "05-contracts"
    evidence_dir = base / "06-evidence"
    letters_dir = base / "07-letters_intelligence"
    outputs_dir = OUTPUTS_ROOT / project["project_folder_name"]

    canonical_letters = load_canonical_letters_payload(project, data_dir)
    if canonical_letters is not None:
        live_inbox_files = list_project_files(letters_dir / "inbox", base, 160)
        historical_inbox_files = canonical_letters["inbox_files"]
        seen_inbox_paths: set[str] = set()
        letter_files = []
        for item in [*live_inbox_files, *historical_inbox_files]:
            relative_path = str(item.get("relative_path") or item.get("name") or "")
            if not relative_path or relative_path.casefold() in seen_inbox_paths:
                continue
            seen_inbox_paths.add(relative_path.casefold())
            letter_files.append(item)
        letter_workbook = canonical_letters["workbook"]
        letter_workspace_tables = merge_canonical_letters_with_inbox(
            project, canonical_letters, letters_dir / "inbox"
        )
        letters_folder = "07-letters_intelligence/inbox + 01-data/import_templates"
        letters_source_detail = (
            "Uses the selected project's 10_letters_intelligence.csv historical register and "
            "automatically evaluates new files in 07-letters_intelligence/inbox/From Contractor "
            "and From Consultant during the local data build."
        )
        letters_detector_status = "Active" if letter_workspace_tables.get("sheets") else "Awaiting Data"
    else:
        letter_files = list_project_files(letters_dir / "inbox", base, 160)
        letter_workbook_path = letters_dir / "letters_intelligence.xlsx"
        letter_workbook = xlsx_summary(letter_workbook_path)
        letter_workspace_tables = build_letters_workspace_tables(project, letter_workbook_path, letters_dir / "inbox")
        letters_folder = "07-letters_intelligence"
        letters_source_detail = "Uses the project-local workbook, inbox ingestion, correspondence links, and issue-thread workflow."
        letters_detector_status = "Active" if letter_workspace_tables.get("sheets") else "Awaiting Data"
    contract_files = list_project_files(contracts_dir / "source", base, 80)
    evidence_files = list_project_files(evidence_dir, base, 80)
    output_files = list_project_files(outputs_dir, OUTPUTS_ROOT, 20)
    controlled_tia = public_controlled_tia_payload(
        build_controlled_tia_snapshot(project), slugify(project["project_folder_name"])
    )
    four_pipeline = build_four_pipeline_snapshot(project, controlled_tia)
    contract_controls = build_contract_controls_snapshot(project, contracts_dir, evidence_dir)
    overview_paths = {
        "projects": source_paths.get("projects", data_dir / "projects.csv"),
        "activities": source_paths.get("activities", data_dir / "activities.csv"),
        "progress_updates": source_paths.get("progress", data_dir / "progress_updates.csv"),
        "evm": source_paths.get("evm", data_dir / "evm.csv"),
        "risks": source_paths.get("risks", data_dir / "risks.csv"),
        "claims": source_paths.get("claims", data_dir / "claims.csv"),
        "contracts": source_paths.get("contracts", data_dir / "contracts.csv"),
        "payments": source_paths.get("payments", data_dir / "payments.csv"),
        "planned_cash_flow": source_paths.get("planned_cash_flow", data_dir / "planned_cash_flow.csv"),
        "milestones": source_paths.get("milestones", data_dir / "milestones.csv"),
        "delay_events": source_paths.get("delay_events", data_dir / "delay_events.csv"),
        "wbs": source_paths.get("wbs", data_dir / "wbs.csv"),
        "s_curve": source_paths.get("s_curve", data_dir / "s_curve.csv"),
    }

    return {
        "overview": {
            "data_sources": {key: len(value) for key, value in rows.items()},
            "source_tables": {
                key: preview_rows_table(rows.get("progress" if key == "progress_updates" else key, []), path) for key, path in overview_paths.items()
            },
            "workspace_tables": {key: workspace_rows_table(rows.get("progress" if key == "progress_updates" else key, []), path) for key, path in overview_paths.items()},
        },
        "letters_intelligence": {
            "folder": letters_folder,
            "inbox_files": letter_files,
            "inbox_file_count": len(letter_files),
            "workbook": letter_workbook,
            "workbook_tables": letter_workspace_tables,
            "detectors": [
                {"name": "Canonical letters source", "status": "Active" if canonical_letters is not None else ("Active" if (letters_dir / "inbox").exists() else "Missing"), "detail": letters_source_detail},
                {"name": "Letters intelligence workflow", "status": letters_detector_status, "detail": "Publishes the selected project's correspondence registers, links, issue threads, and inbox metadata only."},
                {"name": "Project isolation", "status": "Active", "detail": "Only files inside the selected project folder are listed."},
            ],
        },
        "delay_analysis": {
            "folder": "02-delay_analysis",
            "visibility": "workspace",
            "internal_control": {
                "ui_enabled": True,
                "source_scope": "selected_project_only",
                "automatic_engine": "controlled_project_tia",
                "groq_assist": "active_project_evidence_explanation_only",
                "evidence_status": str(controlled_tia.get("status") or "SETUP_REQUIRED"),
                "activation_rule": "Only the active project's approved release may create a controlled run. No other project's XER, event, contract, evidence, or output is used.",
            },
            "logic_mode": "Controlled project-local Time Impact Analysis",
            "controlled_tia": controlled_tia,
            "legacy_status": "Archived and excluded from the active TIA workflow.",
            "detectors": [
                {"name": "Approved release detector", "status": "Active" if controlled_tia.get("source_integrity", {}).get("release_configured") else "Awaiting package", "detail": "Checks only the active project's declared approved TIA release."},
                {"name": "Source integrity control", "status": str(controlled_tia.get("source_integrity", {}).get("signature", {}).get("status") or "not_checked"), "detail": "Validates signed-manifest and source-hash controls without executing the submitted package."},
                {"name": "Run isolation", "status": "Active", "detail": "Automatic drafts and any approved run are written only under this project's 02-delay_analysis/controlled_runs folder."},
            ],
        },
        "four_pipeline": four_pipeline,
        "contract_claims": {
            "folder": "05-contracts",
            "source_files": contract_files,
            "evidence_files": evidence_files,
            "database": sqlite_table_counts(contracts_dir / "contract_claims.db"),
            "knowledge_base": sqlite_table_rows(contracts_dir / "contract_claims.db"),
            "controlled_assessment": contract_controls,
            "clause_library": xlsx_summary(contracts_dir / "source" / "Overall_Contract_clause_library.xlsx"),
            "clause_library_tables": xlsx_workspace_tables(contracts_dir / "source" / "Overall_Contract_clause_library.xlsx"),
            "detectors": [
                {"name": "Contract source detector", "status": "Active" if contract_files else "Missing", "detail": "Finds contract PDFs and clause libraries inside the selected project only."},
                {"name": "Knowledge base detector", "status": "Active" if (contracts_dir / "contract_claims.db").exists() else "Missing", "detail": "Uses the selected project's own SQLite knowledge base."},
                {"name": "Evidence detector", "status": "Active" if evidence_files else "Missing", "detail": "Maps claim evidence files from the selected project evidence folder."},
            ],
        },
        "outputs_and_watchers": {
            "outputs_folder": f"11-outputs/{project['project_folder_name']}",
            "output_files": output_files,
            "watchers": [
                {"name": "Project data detector", "status": "Active", "detail": "Website data generator fingerprints every selected project folder."},
                {"name": "No-Git sync watcher", "status": "Configured" if (ROOT / "RUN_FULL_PROJECT_NO_GIT_SYNC.bat").exists() else "Missing", "detail": "Syncs local code, project folders, generated HTML, and website files to GitHub."},
                {"name": "Generated report watcher", "status": "Configured" if outputs_dir.exists() else "Missing", "detail": "Publishes project-specific HTML, PDF, and PowerPoint outputs from 11-outputs."},
            ],
        },
    }


def pick(row: dict[str, Any], names: list[str]) -> Any:
    lowered = {normalize_field_name(k): v for k, v in row.items()}
    for name in names:
        key = normalize_field_name(name)
        if key in lowered and str(lowered[key]).strip() != "":
            return lowered[key]
    return None


def first_valid(rows: list[dict[str, Any]], names: list[str]) -> Any:
    for row in rows:
        value = pick(row, names)
        if value not in (None, ""):
            return value
    return None


def sum_column(rows: list[dict[str, Any]], names: list[str]) -> float | None:
    total = 0.0
    seen = False
    for row in rows:
        value = safe_float(pick(row, names))
        if value is not None:
            total += value
            seen = True
    return total if seen else None


def choose_measurement(*candidates: tuple[str, float | None]) -> tuple[float | None, str]:
    """Return the first meaningful source while retaining a zero source as a last resort."""
    available: list[tuple[str, float]] = []
    for source, value in candidates:
        if value is None or not math.isfinite(value):
            continue
        available.append((source, value))
        if abs(value) > 1e-9:
            return value, source
    if available:
        return available[0][1], available[0][0]
    return None, "Unavailable"


def weighted_activity_progress(rows: list[dict[str, Any]], progress_fields: list[str]) -> float | None:
    weighted_value = 0.0
    total_weight = 0.0
    for row in rows:
        weight = safe_float(pick(row, ["planned_weight", "weight", "budget_weight"]))
        progress = safe_percent(pick(row, progress_fields))
        if weight is None or progress is None or weight <= 0:
            continue
        weighted_value += weight * progress
        total_weight += weight
    return weighted_value / total_weight if total_weight > 0 else None


def risk_factor(value: Any, *, numeric_kind: str = "rating") -> float | None:
    """Normalize common qualitative or numeric risk values to a 0-1 factor."""
    text = str(value or "").strip().lower()
    if not text:
        return None
    labels = {
        "very low": 0.1,
        "low": 0.25,
        "medium": 0.5,
        "moderate": 0.5,
        "high": 0.75,
        "very high": 1.0,
        "critical": 1.0,
        "severe": 1.0,
        "yes": 0.75,
        "true": 0.75,
        "no": 0.0,
        "false": 0.0,
    }
    if text in labels:
        return labels[text]

    number = safe_float(value)
    if number is None:
        return None
    number = abs(number)
    if numeric_kind == "days":
        if number <= 0:
            return 0.0
        if number <= 7:
            return 0.25
        if number <= 30:
            return 0.5
        if number <= 90:
            return 0.75
        return 1.0
    if number <= 1:
        return number
    if number <= 5:
        return number / 5.0
    if number <= 25:
        return number / 25.0
    return min(number / 100.0, 1.0)


def active_risk(row: dict[str, Any]) -> bool:
    status = str(pick(row, ["status", "risk_status", "current_status"]) or "").strip().lower()
    return status not in {"closed", "resolved", "cancelled", "canceled", "inactive", "withdrawn"}


def qualitative_risk_metrics(rows: list[dict[str, Any]]) -> tuple[float | None, int, str]:
    """Derive an active-risk score from explicit scores or probability/impact fields.

    Closed records are excluded. This is intentionally a transparent management
    indicator, not a replacement for a project-approved risk matrix.
    """
    scores: list[float] = []
    high_count = 0
    high_terms = {"high", "very high", "critical", "severe"}

    for row in rows:
        if not active_risk(row):
            continue

        explicit = risk_factor(pick(row, ["risk_score", "risk rating", "risk_rating", "score", "severity_score"]))
        probability = risk_factor(pick(row, ["probability", "likelihood", "chance"]))
        impact_values = [
            risk_factor(pick(row, ["severity", "impact", "impact_rating", "impact level"])),
            risk_factor(pick(row, ["time_impact_days", "schedule_impact_days", "delay_days"]), numeric_kind="days"),
            risk_factor(pick(row, ["cost_impact", "cost impact", "financial_impact"])),
        ]
        impacts = [value for value in impact_values if value is not None]
        impact = max(impacts) if impacts else None

        if explicit is not None:
            score = explicit * 100.0
        elif probability is not None and impact is not None:
            score = probability * impact * 100.0
        elif probability is not None:
            score = probability * 100.0
        elif impact is not None:
            score = impact * 100.0
        else:
            continue

        scores.append(score)
        qualitative_terms = {
            str(pick(row, ["probability", "likelihood", "chance"]) or "").strip().lower(),
            str(pick(row, ["severity", "impact", "impact_rating", "impact level"]) or "").strip().lower(),
        }
        if score >= 70.0 or qualitative_terms & high_terms:
            high_count += 1

    if scores:
        return average(scores), high_count, "risks.csv:active probability-impact risk matrix"
    return None, high_count, "Unavailable"


def summed_delay_days(rows: list[dict[str, Any]]) -> float | None:
    total = 0.0
    seen = False
    for row in rows:
        value = safe_float(
            pick(
                row,
                [
                    "estimated_delay_days",
                    "Delayed duration after overlap",
                    "Concurrent delay",
                    "delay_days",
                    "Delayed duration",
                ],
            )
        )
        if value is None:
            continue
        total += abs(value)
        seen = True
    return total if seen else None


def average(values: list[float | None]) -> float | None:
    clean = [v for v in values if v is not None and math.isfinite(v)]
    return sum(clean) / len(clean) if clean else None


def health_from_ratio(value: float | None, good: float = 1.0, watch: float = 0.9) -> str:
    if value is None or not math.isfinite(value):
        return "Unknown"
    if value >= good:
        return "Healthy"
    if value >= watch:
        return "Watchlist"
    return "Critical"


def exposure_from_value(value: float | None, medium: float, high: float) -> str:
    amount = value or 0
    if amount >= high:
        return "High"
    if amount >= medium:
        return "Medium"
    return "Low"


def confidence_from_quality(value: float | None) -> str:
    quality = value or 0
    if quality >= 85:
        return "High"
    if quality >= 55:
        return "Medium"
    return "Low"


def priority_rank(value: str) -> int:
    return {"Critical": 0, "High": 1, "Medium": 2, "Watchlist": 3, "Low": 4, "Healthy": 5, "Unknown": 6}.get(value, 6)


def build_decision_reasons(args: dict[str, Any]) -> list[dict[str, str]]:
    reasons: list[dict[str, str]] = []
    project_name = str(args.get("project_display_name") or "Selected project")
    schedule_health = str(args.get("schedule_health") or "Unknown")
    cost_health = str(args.get("cost_health") or "Unknown")
    delay_days = args.get("delay_days") or 0
    claims_exposure = args.get("claims_exposure") or 0
    claimed_days = args.get("claimed_days") or 0
    high_risk_count = args.get("high_risk_count") or 0
    data_confidence = str(args.get("data_confidence") or "Low")
    data_quality = args.get("data_quality") or 0

    if schedule_health == "Critical":
        reasons.append({
            "issue": "Schedule performance below executive threshold",
            "trigger": f"SPI {args.get('spi'):.2f}" if args.get("spi") is not None else "SPI unavailable",
            "impact": "Potential recovery plan and critical path review required.",
            "owner": "Planning Manager",
            "evidence_status": data_confidence,
            "urgency": "High",
            "recommended_action": "Review recovery plan, driving activities, and schedule mitigation before the next management meeting.",
        })
    elif schedule_health == "Watchlist":
        reasons.append({
            "issue": "Schedule performance on watchlist",
            "trigger": f"SPI {args.get('spi'):.2f}" if args.get("spi") is not None else "SPI unavailable",
            "impact": "Schedule trend should be monitored before it becomes critical.",
            "owner": "Planning Manager",
            "evidence_status": data_confidence,
            "urgency": "Medium",
            "recommended_action": "Validate planned versus actual progress and confirm near-critical activities.",
        })

    if cost_health == "Critical":
        reasons.append({
            "issue": "Cost performance below executive threshold",
            "trigger": f"CPI {args.get('cpi'):.2f}" if args.get("cpi") is not None else "CPI unavailable",
            "impact": "Forecast final cost and payment/cost controls may need management intervention.",
            "owner": "Commercial Manager",
            "evidence_status": data_confidence,
            "urgency": "High",
            "recommended_action": "Review cost-control plan, paid/spent records, and EAC exposure.",
        })
    elif cost_health == "Watchlist":
        reasons.append({
            "issue": "Cost performance on watchlist",
            "trigger": f"CPI {args.get('cpi'):.2f}" if args.get("cpi") is not None else "CPI unavailable",
            "impact": "Cost trend should be checked before commitment decisions.",
            "owner": "Commercial Manager",
            "evidence_status": data_confidence,
            "urgency": "Medium",
            "recommended_action": "Validate actual cost and payment records against budget baseline.",
        })

    if delay_days:
        reasons.append({
            "issue": "Delay exposure recorded",
            "trigger": f"{delay_days:.0f} cumulative delay-event days; EOT not yet verified",
            "impact": "Indicative schedule exposure only until critical path, fragnet, and concurrency tests are verified.",
            "owner": "Planning / Claims Team",
            "evidence_status": data_confidence,
            "urgency": "High" if delay_days >= 30 else "Medium",
            "recommended_action": "Review delay causation, critical path status, concurrency, and evidence completeness.",
        })

    if claims_exposure or claimed_days:
        claim_trigger_parts: list[str] = []
        if claims_exposure:
            claim_trigger_parts.append(f"Claim amount EGP {claims_exposure:,.0f}")
        if claimed_days:
            claim_trigger_parts.append(f"Claimed days {claimed_days:.0f}")
        reasons.append({
            "issue": "Claims / EOT exposure available",
            "trigger": " | ".join(claim_trigger_parts),
            "impact": "Commercial strategy and entitlement evidence should be aligned.",
            "owner": "Contracts Manager",
            "evidence_status": data_confidence,
            "urgency": "High" if claims_exposure >= 1000000 else "Medium",
            "recommended_action": "Review entitlement matrix, notice status, and supporting correspondence.",
        })

    if high_risk_count:
        reasons.append({
            "issue": "High-risk records detected",
            "trigger": f"{high_risk_count} high-risk items",
            "impact": "Risk exposure may require mitigation ownership and deadlines.",
            "owner": "Project Manager",
            "evidence_status": data_confidence,
            "urgency": "High",
            "recommended_action": "Assign mitigation owners and confirm closure evidence for high-risk records.",
        })

    if data_quality < 70:
        reasons.append({
            "issue": "Data confidence gap",
            "trigger": f"Data quality {data_quality:.1f}%",
            "impact": "Executive decisions may be based on incomplete source records.",
            "owner": "Project Controls",
            "evidence_status": data_confidence,
            "urgency": "Medium",
            "recommended_action": "Complete missing project controls data before final management decision.",
        })

    if not reasons:
        reasons.append({
            "issue": "No immediate executive trigger",
            "trigger": f"{project_name} is within available decision thresholds",
            "impact": "Continue routine monitoring with the current evidence set.",
            "owner": "Project Manager",
            "evidence_status": data_confidence,
            "urgency": "Low",
            "recommended_action": "Keep source files updated and monitor SPI, CPI, risk, claims, and delay records.",
        })
    return reasons[:5]


def latest_mtime(path: Path) -> str | None:
    if not path.exists():
        return None
    latest = 0.0
    for child in path.rglob("*"):
        if child.is_file():
            latest = max(latest, child.stat().st_mtime)
    if latest == 0:
        return None
    return datetime.fromtimestamp(latest).isoformat(timespec="seconds")


def fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    if not path.exists():
        return ""
    for child in sorted(p for p in path.rglob("*") if p.is_file()):
        rel = child.relative_to(path).as_posix()
        if any(part in {".git", ".venv", "node_modules", "__pycache__"} for part in child.parts):
            continue
        digest.update(rel.encode("utf-8"))
        digest.update(str(child.stat().st_size).encode("ascii"))
        digest.update(str(child.stat().st_mtime_ns).encode("ascii"))
    return digest.hexdigest()


def discover_projects() -> list[dict[str, Any]]:
    """Discover project folders through the shared non-destructive catalog."""
    if not PROJECTS_ROOT.exists():
        return []
    if str(ROOT) not in sys.path:
        sys.path.insert(0, str(ROOT))
    from src.construction_system.project_catalog import (
        discover_projects as catalog_discover_projects,
        ensure_project_structure,
    )

    template_dir = PROJECTS_ROOT / "_PROJECT_TEMPLATE"
    if template_dir.exists():
        ensure_project_structure(template_dir)

    projects: list[dict[str, Any]] = []
    for record in catalog_discover_projects(PROJECTS_ROOT):
        project_dir = Path(str(record["project_dir"]))
        project_json = read_json(project_dir / "project.json")
        project_id = str(record["project_id"])
        projects.append(
            {
                "project_id": project_id,
                "project_key": slugify(project_id).lower(),
                "project_folder_name": project_dir.name,
                "project_display_name": str(record.get("project_display_name") or project_dir.name),
                "sector": str(record.get("sector_name") or "Unassigned"),
                "meeting_url": (
                    project_json.get("meeting_url")
                    or project_json.get("conference_url")
                    or project_json.get("teams_url")
                    or project_json.get("zoom_url")
                    or project_json.get("google_meet_url")
                ),
                "path": project_dir,
            }
        )
    return projects


def build_project_record(project: dict[str, Any]) -> dict[str, Any]:
    base = Path(project["path"])
    data_dir = base / "01-data" / "import_templates"
    rows, source_paths = load_project_input_rows(data_dir)


    project_meta = rows["projects"][0] if rows["projects"] else {}

    # Project summary records are the authoritative source for project identity and progress.
    # EVM records are activity-level, so every monetary EVM metric is summed across the
    # selected project's complete file, matching the legacy Streamlit calculation flow.
    project_contract_value = safe_float(first_valid(rows["projects"], ["contract_value", "budget", "bac"]))
    contract_register_value = sum_column(rows["contracts"], ["contract_value", "original_value", "value", "amount", "bac"])
    evm_bac_total = sum_column(rows["evm"], ["bac", "budget_at_completion"])
    contract_value, contract_source = choose_measurement(
        ("projects.csv:contract_value", project_contract_value),
        ("contracts.csv:original_value total", contract_register_value),
        ("evm.csv:BAC total", evm_bac_total),
    )

    paid_from_payments = sum_column(rows["payments"], ["paid_amount", "paid amount", "paid", "payment_amount", "amount"])
    paid_from_contracts = sum_column(rows["contracts"], ["paid_to_date", "paid amount", "paid_amount"])
    paid_amount, paid_source = choose_measurement(
        ("payments.csv:paid amount total", paid_from_payments),
        ("contracts.csv:paid_to_date total", paid_from_contracts),
    )

    evm_pv_total = sum_column(rows["evm"], ["pv", "planned_value", "planned value"])
    evm_ev_total = sum_column(rows["evm"], ["ev", "earned_value", "earned value"])
    evm_ac_total = sum_column(rows["evm"], ["ac", "actual_cost", "actual cost"])
    progress_pv_total = sum_column(rows["progress"], ["planned_value", "planned value"])
    progress_ev_total = sum_column(rows["progress"], ["earned_value", "earned value"])
    progress_ac_total = sum_column(rows["progress"], ["actual_cost", "actual cost"])

    project_planned_progress = safe_percent(
        first_valid(rows["projects"], ["planned_progress_percent", "planned_progress", "planned percent", "baseline_progress"])
    )
    project_actual_progress = safe_percent(
        first_valid(rows["projects"], ["actual_progress_percent", "actual_progress", "overall_progress", "progress"])
    )
    activity_planned_progress = weighted_activity_progress(rows["activities"], ["planned_progress", "planned percent"])
    activity_actual_progress = weighted_activity_progress(rows["activities"], ["actual_progress", "actual percent", "progress"])

    bac, bac_source = choose_measurement(
        ("evm.csv:BAC total", evm_bac_total),
        ("projects.csv:contract_value", contract_value),
    )
    derived_pv = bac * project_planned_progress if bac is not None and project_planned_progress is not None else None
    derived_ev = bac * project_actual_progress if bac is not None and project_actual_progress is not None else None
    pv, pv_source = choose_measurement(
        ("evm.csv:PV total", evm_pv_total),
        ("progress_updates.csv:planned_value total", progress_pv_total),
        ("projects.csv:contract value x planned progress", derived_pv),
    )
    ev, ev_source = choose_measurement(
        ("evm.csv:EV total", evm_ev_total),
        ("progress_updates.csv:earned_value total", progress_ev_total),
        ("projects.csv:contract value x actual progress", derived_ev),
    )
    ac, ac_source = choose_measurement(
        ("evm.csv:AC total", evm_ac_total),
        ("progress_updates.csv:actual_cost total", progress_ac_total),
        ("payments.csv:paid amount total (cost proxy)", paid_amount),
    )
    spent_amount = ac

    planned_progress, planned_progress_source = choose_measurement(
        ("projects.csv:planned_progress_percent", project_planned_progress),
        ("evm.csv:PV / BAC", pv / bac if pv is not None and bac not in (None, 0) else None),
        ("activities.csv:weighted planned progress", activity_planned_progress),
    )
    actual_progress, actual_progress_source = choose_measurement(
        ("projects.csv:actual_progress_percent", project_actual_progress),
        ("evm.csv:EV / BAC", ev / bac if ev is not None and bac not in (None, 0) else None),
        ("activities.csv:weighted actual progress", activity_actual_progress),
    )

    spi = ev / pv if ev is not None and pv not in (None, 0) else None
    cpi = ev / ac if ev is not None and ac not in (None, 0) else None
    sv = ev - pv if ev is not None and pv is not None else None
    cv = ev - ac if ev is not None and ac is not None else None
    eac = bac / cpi if bac is not None and cpi not in (None, 0) else None
    etc = eac - ac if eac is not None and ac is not None else None
    vac = bac - eac if bac is not None and eac is not None else None
    remaining_value = contract_value - paid_amount if contract_value is not None and paid_amount is not None else None

    risk_score, high_risk_count, risk_source = qualitative_risk_metrics(rows["risks"])
    delay_days = summed_delay_days(rows["delay_events"])
    claims_exposure = sum_column(rows["claims"], ["claim_amount", "claimed_amount", "amount", "eot_exposure", "exposure"])
    claimed_days = sum_column(rows["claims"], ["claimed_days", "claim_days", "eot_days", "claimed duration"])
    status = str(first_valid(rows["projects"], ["status", "project_status"]) or "Active")

    data_quality_fields = [
        contract_value,
        paid_amount,
        planned_progress,
        actual_progress,
        spi,
        cpi,
        risk_score,
    ]
    metric_completeness = sum(value is not None for value in data_quality_fields) / len(data_quality_fields)
    required_source_sets = ("projects", "activities", "evm", "risks")
    source_completeness = sum(bool(rows[name]) for name in required_source_sets) / len(required_source_sets)
    data_quality = round((metric_completeness * 0.75 + source_completeness * 0.25) * 100, 1)

    if (spi is not None and spi < 0.9) or (cpi is not None and cpi < 0.9) or high_risk_count > 0:
        decision_required = True
    else:
        decision_required = bool(delay_days or claims_exposure or claimed_days)

    schedule_health = health_from_ratio(spi)
    cost_health = health_from_ratio(cpi)
    delay_exposure = exposure_from_value(delay_days, medium=1, high=30)
    claim_exposure_level = exposure_from_value(claims_exposure, medium=1, high=1000000)
    data_confidence = confidence_from_quality(data_quality)
    advanced_analytics = build_advanced_analytics(
        project_key=str(project["project_key"]),
        rows=rows,
        contract_value=contract_value,
        output_dir=DATA_ROOT / "analytics",
    )
    priority_inputs = [
        "High" if decision_required else "Low",
        "High" if schedule_health == "Critical" else schedule_health,
        "High" if cost_health == "Critical" else cost_health,
        delay_exposure,
        claim_exposure_level,
        "High" if high_risk_count else "Low",
        "High" if data_confidence == "Low" else data_confidence,
    ]
    decision_priority = sorted(priority_inputs, key=priority_rank)[0]
    decision_reasons = build_decision_reasons({
        "project_display_name": project["project_display_name"],
        "schedule_health": schedule_health,
        "cost_health": cost_health,
        "spi": spi,
        "cpi": cpi,
        "delay_days": delay_days,
        "delay_assessment": "Indicative schedule exposure only. Verify critical path, fragnet logic, and concurrency in Primavera P6 before using as EOT.",
        "claims_exposure": claims_exposure,
        "claimed_days": claimed_days,
        "high_risk_count": high_risk_count,
        "data_confidence": data_confidence,
        "data_quality": data_quality,
        "data_quality_components": {
            "metric_completeness": round(metric_completeness * 100, 1),
            "required_source_completeness": round(source_completeness * 100, 1),
            "required_source_sets": list(required_source_sets),
        },
        "advanced_analytics": advanced_analytics,
    })

    features = build_feature_payload(project, rows, source_paths)
    chart_payloads = build_project_chart_payloads(
        project_id=str(project["project_id"]),
        project_key=str(project["project_key"]),
        data_dir=data_dir,
        vercel_dir=base / "vercel",
        delay_dir=base / "02-delay_analysis" / "unified_tia_csv",
        payment_rows=rows["payments"],
        delay_event_rows=rows["delay_events"],
        activity_rows=rows["activities"],
        read_csv_rows=read_csv_rows,
        workspace_rows=rows,
        project_metrics={
            "bac": bac,
            "pv": pv,
            "ev": ev,
            "ac": ac,
            "eac": eac,
            "planned_progress": planned_progress,
            "actual_progress": actual_progress,
            "spi": spi,
            "cpi": cpi,
        },
    )

    return {
        **{k: v for k, v in project.items() if k != "path"},
        "status": status,
        "contract_value": contract_value,
        "paid_amount": paid_amount,
        "spent_amount": spent_amount,
        "remaining_value": remaining_value,
        "planned_progress": planned_progress,
        "actual_progress": actual_progress,
        "progress_variance": actual_progress - planned_progress if actual_progress is not None and planned_progress is not None else None,
        "bac": bac,
        "pv": pv,
        "ev": ev,
        "ac": ac,
        "sv": sv,
        "cv": cv,
        "eac": eac,
        "etc": etc,
        "vac": vac,
        "spi": spi,
        "cpi": cpi,
        "risk_score": risk_score,
        "high_risk_count": high_risk_count,
        "delay_days": delay_days,
        "delay_event_count": len(rows["delay_events"]),
        "claims_exposure": claims_exposure,
        "claimed_days": claimed_days,
        "risk_record_count": len(rows["risks"]),
        "planned_start": first_valid(rows["projects"], ["planned_start", "project_start", "baseline_start"]),
        "planned_finish": first_valid(rows["projects"], ["planned_finish", "project_finish", "baseline_finish"]),
        "forecast_finish": first_valid(rows["projects"], ["forecast_finish", "current_finish", "forecast date"]),
        "schedule_health": schedule_health,
        "cost_health": cost_health,
        "delay_exposure": delay_exposure,
        "claim_exposure_level": claim_exposure_level,
        "data_confidence": data_confidence,
        "decision_priority": decision_priority,
        "decision_reasons": decision_reasons,
        "activity_count": len(rows["activities"]),
        "milestone_count": len(rows["milestones"]),
        "data_quality": data_quality,
        "advanced_analytics": advanced_analytics,
        "decision_required": decision_required,
        "last_updated": latest_mtime(base),
        "fingerprint": fingerprint(base),
        "source_files": {
            key: len(value) for key, value in rows.items()
        },
        "metric_sources": {
            "contract_value": {"source": contract_source, "aggregation": "project summary or project-level fallback"},
            "paid_amount": {"source": paid_source, "aggregation": "sum of selected project records"},
            "spent_amount": {"source": ac_source, "aggregation": "sum of selected project records"},
            "planned_progress": {"source": planned_progress_source, "aggregation": "project summary or project-level calculation"},
            "actual_progress": {"source": actual_progress_source, "aggregation": "project summary or project-level calculation"},
            "bac": {"source": bac_source, "aggregation": "sum of selected project EVM records"},
            "pv": {"source": pv_source, "aggregation": "sum of selected project EVM records"},
            "ev": {"source": ev_source, "aggregation": "sum of selected project EVM records"},
            "ac": {"source": ac_source, "aggregation": "sum of selected project EVM records"},
            "risk_score": {"source": risk_source, "aggregation": "project risk register"},
            "delay_days": {"source": "delay_events.csv:estimated or overlap-adjusted duration", "aggregation": "cumulative event exposure; not a verified EOT calculation"},
            "claims_exposure": {"source": "claims.csv:claimed_amount", "aggregation": "sum of selected project claim records"},
        },
        "features": features,
        "chart_payloads": chart_payloads,
        "reports": {
            "executive_dashboard": f"/generated/{slugify(project['project_folder_name'])}/01_executive_dashboard.html",
            "master_dashboard": f"/generated/{slugify(project['project_folder_name'])}/02_master_dashboard.html",
            "elite_svg_charts": f"/generated/{slugify(project['project_folder_name'])}/03_elite_svg_charts.html",
            "linked_executive_dashboard": f"/generated/{slugify(project['project_folder_name'])}/04_linked_executive_dashboard.html",
        },
    }


def copy_if_changed(source: Path, target: Path) -> None:
    if target.exists() and target.stat().st_size == source.stat().st_size:
        try:
            if _sha256_file(target) == _sha256_file(source):
                return
        except OSError:
            pass
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)


def _json_default(value: Any) -> str:
    """Keep generated payloads portable when local services return Path metadata."""

    if isinstance(value, Path):
        return str(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def write_json_if_changed(path: Path, payload: dict[str, Any]) -> bool:
    """Write generated data only when its content changes, preserving project freshness."""
    content = json.dumps(payload, indent=2, ensure_ascii=False, default=_json_default) + "\n"
    if path.exists() and path.read_text(encoding="utf-8") == content:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")
    return True


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def copy_generated_outputs(projects: list[dict[str, Any]]) -> None:
    """Publish selected-project report artifacts without rewriting unchanged projects."""
    GENERATED_ROOT.mkdir(parents=True, exist_ok=True)
    OUTPUTS_ROOT.mkdir(parents=True, exist_ok=True)
    active_public_dirs = {slugify(project["project_folder_name"]) for project in projects}
    for orphan in GENERATED_ROOT.iterdir():
        if orphan.is_dir() and orphan.name not in active_public_dirs:
            shutil.rmtree(orphan)

    for project in projects:
        project_folder = project["project_folder_name"]
        source = SOURCE_OUTPUTS_ROOT / project_folder
        output_dir = OUTPUTS_ROOT / project_folder
        output_dir.mkdir(parents=True, exist_ok=True)
        if source.exists() and source.resolve() != output_dir.resolve():
            for html_report in sorted(source.glob("*.html")):
                copy_if_changed(html_report, output_dir / html_report.name)

        public_slug = slugify(project_folder)
        artifacts = ensure_project_report_artifacts(project, output_dir, public_slug=public_slug)
        project["report_package"] = artifacts.pop("published_project_package", None)
        project["report_artifacts"] = artifacts

        # The Universal Report Engine is a local controlled tool.  Its catalogue
        # is always generated for the owning project, while report packages are
        # published only after a local, source-bound engine run creates them.
        project_dir = PROJECTS_ROOT / str(project.get("sector") or "") / str(project_folder)
        if project_dir.is_dir():
            project["universal_report_engine"] = ensure_universal_report_engine_catalog(
                project, project_dir, output_dir, public_slug
            )
        project["features"]["outputs_and_watchers"]["output_files"] = list_project_files(output_dir, OUTPUTS_ROOT, 80)

        target = GENERATED_ROOT / public_slug
        target.mkdir(parents=True, exist_ok=True)
        for artifact in sorted(output_dir.iterdir()):
            if artifact.is_file() and artifact.suffix.lower() in {".html", ".pdf", ".pptx", ".docx", ".zip"}:
                copy_if_changed(artifact, target / artifact.name)

        # Publish only release-approved Universal Report Engine artifacts.
        # Context, raw source inventory, validation internals, and draft/failed
        # packages stay on the local workstation under the selected project's
        # output folder.
        universal_root = output_dir / "universal-report-engine"
        if universal_root.is_dir():
            for manifest_path in sorted(universal_root.rglob(FAMILY_MANIFEST_NAME)):
                manifest = read_json(manifest_path)
                if not isinstance(manifest, dict) or not is_released_artifact(manifest):
                    continue
                if manifest.get("project_id") != project.get("project_id"):
                    continue
                if manifest.get("project_key") != project.get("project_key"):
                    continue
                for relative_path in (manifest.get("artifacts") or {}).values():
                    if not isinstance(relative_path, str):
                        continue
                    artifact = output_dir / relative_path
                    if artifact.is_file() and artifact.suffix.lower() in {".html", ".pdf", ".pptx", ".zip"}:
                        copy_if_changed(artifact, target / relative_path)


def copy_controlled_tia_exhibits(projects: list[dict[str, Any]]) -> None:
    """Publish only the active project's declared controlled TIA exhibits.

    The submission manifest supplies the relative path for each exhibit. The
    path is resolved below the owning project's approved release folder and
    the browser payload receives a static public URL, never a workstation path.
    """
    for project in projects:
        delay_analysis = project.get("features", {}).get("delay_analysis", {})
        controlled = delay_analysis.get("controlled_tia", {})
        event_section = controlled.get("events_and_fragnets", {}) if isinstance(controlled, dict) else {}
        event_exhibits = event_section.get("event_exhibits", []) if isinstance(event_section, dict) else []
        view_exhibits = controlled.get("view_exhibits", []) if isinstance(controlled, dict) else []
        if not isinstance(event_exhibits, list):
            event_exhibits = []
        if not isinstance(view_exhibits, list):
            view_exhibits = []
        if not event_exhibits and not view_exhibits:
            continue

        # Browser records intentionally omit local paths. Reconstruct the
        # owning workspace from the discovered sector/folder identity instead
        # of using a hidden fallback or another project's release folder.
        project_dir = PROJECTS_ROOT / str(project.get("sector") or "") / str(project.get("project_folder_name") or "")
        if not project_dir.is_dir():
            continue
        manifest = read_json(project_dir / "project_manifest.json")
        release = manifest.get("approved_tia_release") if isinstance(manifest.get("approved_tia_release"), dict) else {}
        source_path = Path(str(release.get("source_path") or ""))
        source_dir = source_path if source_path.is_absolute() else (project_dir / source_path)
        try:
            source_dir = source_dir.resolve()
            source_dir.relative_to(project_dir.resolve())
        except (OSError, ValueError):
            continue

        for exhibit_set, target_name in (
            (event_exhibits, "tia-controlled-event-exhibits"),
            (view_exhibits, "tia-controlled-view-exhibits"),
        ):
            target = GENERATED_ROOT / slugify(project["project_folder_name"]) / target_name
            for exhibit in exhibit_set:
                if not isinstance(exhibit, dict):
                    continue
                relative_path = str(exhibit.get("source_relative_path") or "").strip()
                if not relative_path:
                    continue
                try:
                    source = (source_dir / relative_path).resolve()
                    source.relative_to(source_dir)
                except (OSError, ValueError):
                    continue
                if source.exists() and source.is_file():
                    public_name = f"{slugify(source.stem)}{source.suffix.lower()}"
                    copy_if_changed(source, target / public_name)
                # The file name remains traceable, but the filesystem-relative path
                # is unnecessary after publication and is removed from browser data.
                exhibit.pop("source_relative_path", None)


def copy_legacy_tia_submitted_assets_archived(projects: list[dict[str, Any]]) -> None:
    for project in projects:
        delay_analysis = project.get("features", {}).get("delay_analysis", {})
        submitted = delay_analysis.get("submitted_tia", {})
        if not submitted.get("available"):
            submitted = {}
        guide_root = Path(str(submitted.get("guide_folder") or ""))
        if guide_root.exists():
            target = GENERATED_ROOT / slugify(project["project_folder_name"]) / "tia-submitted-guide"
            target.mkdir(parents=True, exist_ok=True)
            for visual in submitted.get("visuals", []):
                source = guide_root / str(visual.get("relative_path", ""))
                if source.exists() and source.is_file():
                    target_name = f"{slugify(source.stem)}{source.suffix.lower()}"
                    copy_if_changed(source, target / target_name)

        submitted_visuals = delay_analysis.get("submitted_visuals", {})
        visual_root = PROJECTS_ROOT / project["sector"] / project["project_folder_name"] / "02-delay_analysis" / "submitted_visuals"
        if not submitted_visuals.get("available") or not visual_root.exists():
            continue
        target = GENERATED_ROOT / slugify(project["project_folder_name"]) / "tia-submitted-exhibits"
        target.mkdir(parents=True, exist_ok=True)
        for visual in submitted_visuals.get("visuals", []):
            source = visual_root / str(visual.get("relative_path", ""))
            if source.exists() and source.is_file():
                target_name = f"{slugify(source.stem)}{source.suffix.lower()}"
                copy_if_changed(source, target / target_name)


def build_portfolio_decision_brief(projects: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for project in projects:
        reasons = project.get("decision_reasons") or []
        for index, reason in enumerate(reasons[:2], start=1):
            urgency = str(reason.get("urgency") or "Low")
            if urgency == "Low" and not project.get("decision_required"):
                continue
            items.append({
                "decision_id": f"DEC-{project['project_key']}-{index:02d}",
                "project_key": project["project_key"],
                "project_id": project["project_id"],
                "project_display_name": project["project_display_name"],
                "project_folder_name": project["project_folder_name"],
                "sector": project["sector"],
                "priority": project.get("decision_priority") or urgency,
                "issue": reason.get("issue") or "Management decision required",
                "trigger": reason.get("trigger") or "Project control threshold triggered",
                "impact": reason.get("impact") or "Potential management impact.",
                "owner": reason.get("owner") or "Project Manager",
                "evidence_status": reason.get("evidence_status") or project.get("data_confidence") or "Low",
                "urgency": urgency,
                "recommended_action": reason.get("recommended_action") or "Review source evidence and assign owner.",
                "last_updated": project.get("last_updated"),
            })
    return sorted(
        items,
        key=lambda item: (
            priority_rank(str(item.get("urgency") or "Low")),
            priority_rank(str(item.get("priority") or "Low")),
            item.get("project_display_name") or "",
        ),
    )[:18]


def portfolio_project_summary(project: dict[str, Any]) -> dict[str, Any]:
    """Keep portfolio JSON executive-sized; detailed evidence remains per project."""
    fields = (
        "project_id",
        "project_key",
        "project_folder_name",
        "project_display_name",
        "sector",
        "status",
        "contract_value",
        "paid_amount",
        "spent_amount",
        "remaining_value",
        "planned_progress",
        "actual_progress",
        "progress_variance",
        "bac",
        "pv",
        "ev",
        "ac",
        "sv",
        "cv",
        "eac",
        "etc",
        "vac",
        "spi",
        "cpi",
        "risk_score",
        "high_risk_count",
        "risk_record_count",
        "delay_days",
        "delay_event_count",
        "claims_exposure",
        "claimed_days",
        "planned_start",
        "planned_finish",
        "forecast_finish",
        "schedule_health",
        "cost_health",
        "delay_exposure",
        "claim_exposure_level",
        "data_confidence",
        "decision_priority",
        "decision_reasons",
        "activity_count",
        "milestone_count",
        "data_quality",
        "decision_required",
        "last_updated",
        "source_files",
        "metric_sources",
        "reports",
    )
    return {field: project.get(field) for field in fields}


def build_portfolio(projects: list[dict[str, Any]]) -> dict[str, Any]:
    total_contract = sum(p["contract_value"] or 0 for p in projects)
    total_paid = sum(p["paid_amount"] or 0 for p in projects)
    total_spent = sum(p["spent_amount"] or 0 for p in projects)
    total_remaining = sum(p["remaining_value"] or 0 for p in projects)
    progress_values = [p["actual_progress"] for p in projects]
    weighted_progress = None
    weighted_basis = sum(p["contract_value"] or 0 for p in projects if p["actual_progress"] is not None)
    if weighted_basis > 0:
        weighted_progress = sum((p["contract_value"] or 0) * (p["actual_progress"] or 0) for p in projects) / weighted_basis

    sectors: dict[str, dict[str, Any]] = {}
    for project in projects:
        sector = sectors.setdefault(
            project["sector"],
            {
                "sector": project["sector"],
                "project_count": 0,
                "contract_value": 0,
                "paid_amount": 0,
                "spent_amount": 0,
                "average_progress": None,
                "average_spi": None,
                "average_cpi": None,
                "average_risk_score": None,
                "delayed_projects": 0,
                "decisions_required": 0,
            },
        )
        sector["project_count"] += 1
        sector["contract_value"] += project["contract_value"] or 0
        sector["paid_amount"] += project["paid_amount"] or 0
        sector["spent_amount"] += project["spent_amount"] or 0
        sector["delayed_projects"] += 1 if (project["delay_days"] or 0) > 0 or (project["spi"] is not None and project["spi"] < 1) else 0
        sector["decisions_required"] += 1 if project["decision_required"] else 0

    for sector in sectors.values():
        sector_projects = [p for p in projects if p["sector"] == sector["sector"]]
        sector["average_progress"] = average([p["actual_progress"] for p in sector_projects])
        sector["average_spi"] = average([p["spi"] for p in sector_projects])
        sector["average_cpi"] = average([p["cpi"] for p in sector_projects])
        sector["average_risk_score"] = average([p["risk_score"] for p in sector_projects])
        sector["average_data_quality"] = average([p["data_quality"] for p in sector_projects])
        sector["critical_schedule_projects"] = sum(1 for p in sector_projects if p.get("schedule_health") == "Critical")
        sector["critical_cost_projects"] = sum(1 for p in sector_projects if p.get("cost_health") == "Critical")
        sector["high_priority_projects"] = sum(1 for p in sector_projects if p.get("decision_priority") == "High")

    warning_summary = {
        "schedule_critical": sum(1 for p in projects if p.get("schedule_health") == "Critical"),
        "cost_critical": sum(1 for p in projects if p.get("cost_health") == "Critical"),
        "delay_high": sum(1 for p in projects if p.get("delay_exposure") == "High"),
        "claims_high": sum(1 for p in projects if p.get("claim_exposure_level") == "High"),
        "low_confidence": sum(1 for p in projects if p.get("data_confidence") == "Low"),
        "high_priority": sum(1 for p in projects if p.get("decision_priority") == "High"),
    }

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "project_count": len(projects),
        "sector_count": len(sectors),
        "totals": {
            "contract_value": total_contract,
            "paid_amount": total_paid,
            "spent_amount": total_spent,
            "remaining_value": total_remaining,
            "average_progress": weighted_progress if weighted_progress is not None else average(progress_values),
            "average_spi": average([p["spi"] for p in projects]),
            "average_cpi": average([p["cpi"] for p in projects]),
            "average_risk_score": average([p["risk_score"] for p in projects]),
            "delayed_projects": sum(1 for p in projects if (p["delay_days"] or 0) > 0 or (p["spi"] is not None and p["spi"] < 1)),
            "high_risk_projects": sum(1 for p in projects if (p["risk_score"] or 0) >= 70 or p["high_risk_count"] > 0),
            "claims_exposure": sum(p["claims_exposure"] or 0 for p in projects),
            "decisions_required": sum(1 for p in projects if p["decision_required"]),
            "average_data_quality": average([p["data_quality"] for p in projects]),
        },
        "warning_summary": warning_summary,
        "decision_brief": build_portfolio_decision_brief(projects),
        "sectors": sorted(sectors.values(), key=lambda item: item["sector"]),
        "projects": sorted(
            (portfolio_project_summary(project) for project in projects),
            key=lambda item: (item["sector"], item["project_display_name"]),
        ),
    }


def _generate() -> None:
    DATA_ROOT.mkdir(parents=True, exist_ok=True)
    WEBSITE_SOURCE_GENERATED.mkdir(parents=True, exist_ok=True)
    raw_projects = discover_projects()
    project_records = [build_project_record(project) for project in raw_projects]
    copy_generated_outputs(project_records)
    copy_controlled_tia_exhibits(project_records)
    # Historic submitted-guide assets remain recoverable on disk but are not
    # copied into active website payloads or reports.  Controlled TIA runs are
    # published through each project's approved source contract instead.
    portfolio = build_portfolio(project_records)
    write_json_if_changed(DATA_ROOT / "portfolio.json", portfolio)
    projects_dir = DATA_ROOT / "projects"
    projects_dir.mkdir(parents=True, exist_ok=True)
    active_project_files = {f"{project['project_key']}.json" for project in project_records}
    for project in project_records:
        write_json_if_changed(projects_dir / f"{project['project_key']}.json", project)
    for stale in projects_dir.glob("*.json"):
        if stale.name not in active_project_files:
            stale.unlink()
    # The website loads the selected project JSON on demand and validates its
    # project_id/project_key before rendering. Remove the old aggregate module
    # payload so the repository does not carry a large duplicate of all projects.
    legacy_workspace_payload = WEBSITE_SOURCE_GENERATED / "project-workspace-payloads.json"
    if legacy_workspace_payload.exists():
        legacy_workspace_payload.unlink()
    try:
        from pih_data_guardrails import run_guardrails

        block_on_issues = "--block-on-guardrails" in sys.argv
        guardrails = run_guardrails(
            portfolio_json_path=DATA_ROOT / "portfolio.json",
            projects_json_dir=projects_dir,
            projects_root=PROJECTS_ROOT,
            backup_dir=ROOT / "12-logs" / "_guardrail_backups",
            action_db_path=ROOT / "12-logs" / "actions.db",
            report_path=ROOT / "12-logs" / "guardrail_report_latest.md",
            block_on_issues=block_on_issues,
        )
        portfolio["guardrails"] = guardrails
        write_json_if_changed(DATA_ROOT / "portfolio.json", portfolio)
        print(
            "Guardrails: "
            f"{guardrails['status']} "
            f"({guardrails['block_count']} block, {guardrails['warn_count']} warn)."
        )
        if block_on_issues and not guardrails["ok"]:
            raise SystemExit(1)
    except SystemExit:
        raise
    except Exception as exc:
        portfolio["guardrails"] = {
            "status": "Error",
            "mode": "WARN",
            "ok": True,
            "block_count": 0,
            "warn_count": 1,
            "issue_count": 1,
            "report_path": "12-logs/guardrail_report_latest.md",
            "snapshot_dir": None,
            "last_checked": datetime.now().isoformat(timespec="seconds"),
            "top_issues": [
                {
                    "severity": "WARN",
                    "effective_severity": "WARN",
                    "scope": "pipeline",
                    "project_id": "portfolio",
                    "project_key": "portfolio",
                    "project_display_name": "Portfolio",
                    "field": "guardrail_runtime_error",
                    "message": str(exc),
                }
            ],
        }
        write_json_if_changed(DATA_ROOT / "portfolio.json", portfolio)
        print(f"Guardrails: Error logged without blocking build: {exc}")
    print(f"Generated Next.js website data for {len(project_records)} projects.")


def main() -> None:
    """Prevent synchronizers from observing a partially regenerated payload set."""
    lock_path = ROOT / ".sync_state" / "generator.lock"
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    lock_path.write_text(
        json.dumps({"pid": os.getpid(), "started_at": datetime.now().isoformat(timespec="seconds")}),
        encoding="utf-8",
    )
    try:
        _generate()
    finally:
        try:
            lock_path.unlink(missing_ok=True)
        except OSError:
            pass


if __name__ == "__main__":
    main()
